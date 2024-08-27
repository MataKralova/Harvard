from openpyxl import load_workbook
from operator import itemgetter
import tkinter as tk
from tkinter import ttk


def main():
    # Load the entire workbook
    wb = load_workbook("Food Data.xlsx")

    ws = wb["Values"]
    all_rows = list(ws.rows)

    # Get the first row as keys
    keys = [cell.value for cell in ws[1]]

    # Create a list of dictionaries
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(keys, row))
        data.append(row_dict)

    ####################### WIDGET #####################################

    def display_selection():
        selected_1 = nutrients_1.get()
        selected_2 = nutrients_2.get()
        selected_3 = nutrients_3.get()
        selected_4 = nutrients_4.get()
        selected_5 = nutrients_5.get()
        selected_6 = nutrients_6.get()

        # Create a list sorted from biggest by values in main selected column
        new_list = sorted(data[0:175], key=itemgetter(selected_1), reverse=True)
        
        # Populate result values for column "Potravina (100 g)"
        for row in range (0,10):
            foods[row].set(new_list[row]["Potravina (100 g)"])

        # Populate result values for column "Jedlo (obvyklá porcia)"
        for row in range (0,10):
            sizes[row].set(new_list[row]["Jedlo (obvyklá porcia)"])

        # Populate result values for main selected quantities column
        for row in range (0,10):
            quantities_1[row].set(new_list[row][selected_1])

        # Populate result values for 2nd selected quantities column
        for row in range (0,10):
            quantities_2[row].set(new_list[row][selected_2])

        # Populate result values for 3rd selected quantities column
        for row in range (0,10):
            quantities_3[row].set(new_list[row][selected_3])

        # Populate result values for 4th selected quantities column
        for row in range (0,10):
            quantities_4[row].set(new_list[row][selected_4])

        # Populate result values for 5th selected quantities column
        for row in range (0,10):
            quantities_5[row].set(new_list[row][selected_5])

        # Populate result values for 6th selected quantities column
        for row in range (0,10):
            quantities_6[row].set(new_list[row][selected_6])

    # Create tkinter window
    window = tk.Tk(className=" Food Nutritional Values")
    window.geometry("1500x305")

    # Define labels for the lists
    label = tk.Label(window, text=" Select nutrients:").grid(column=1, row=15, padx=10, pady=25)

    # Create display values for comboboxes
    n1 = tk.StringVar()
    n2 = tk.StringVar()
    n3 = tk.StringVar()
    n4 = tk.StringVar()
    n5 = tk.StringVar()
    n6 = tk.StringVar()

    # Create comboboxes
    nutrients_1 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n1)
    nutrients_2 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n2)
    nutrients_3 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n3)
    nutrients_4 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n4)
    nutrients_5 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n5)
    nutrients_6 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n6)

    # Add combobox dropdown lists
    values = []
    for key in keys:
        values.append(key)
    nutrients_1["values"] = values
    nutrients_2["values"] = values
    nutrients_3["values"] = values
    nutrients_4["values"] = values
    nutrients_5["values"] = values
    nutrients_6["values"] = values

    # Show comboboxes with default values
    nutrients_1.grid(column=2, row=15)
    nutrients_1.current(0)
    nutrients_2.grid(column=3, row=15)
    nutrients_2.current(0)
    nutrients_3.grid(column=4, row=15)
    nutrients_3.current(0)
    nutrients_4.grid(column=5, row=15)
    nutrients_4.current(0)
    nutrients_5.grid(column=6, row=15)
    nutrients_5.current(0)
    nutrients_6.grid(column=7, row=15)
    nutrients_6.current(0)

    # Create StringVar lists to associate with labels
    foods = []
    sizes = []
    quantities_1 = []
    quantities_2 = []
    quantities_3 = []
    quantities_4 = []
    quantities_5 = []
    quantities_6 = []

    # Create result labels
    for row in range (0,10):
        # Column 0 showing foods
        food = tk.StringVar()
        food.set("                              ")
        foods.append(food)
        label = tk.Label(window, textvariable=foods[row]).grid(column=0, row=row+100, padx=25)
        # Column 1 showing sizes
        size = tk.StringVar()
        size.set("               ")
        sizes.append(size)
        label = tk.Label(window, textvariable=sizes[row]).grid(column=1, row=row+100)
        # Column 2 showing main selected quantities
        quantity_1 = tk.StringVar()
        quantity_1.set("                         ")
        quantities_1.append(quantity_1)           
        label = tk.Label(window, textvariable=quantities_1[row]).grid(column=2, row=row+100)
        # Column 3 showing 2nd selected quantities
        quantity_2 = tk.StringVar()
        quantity_2.set("                         ")
        quantities_2.append(quantity_2)           
        label = tk.Label(window, textvariable=quantities_2[row]).grid(column=3, row=row+100)
        # Column 4 showing 3rd selected quantities
        quantity_3 = tk.StringVar()
        quantity_3.set("                         ")
        quantities_3.append(quantity_3)           
        label = tk.Label(window, textvariable=quantities_3[row]).grid(column=4, row=row+100)
        # Column 5 showing 4th selected quantities
        quantity_4 = tk.StringVar()
        quantity_4.set("                         ")
        quantities_4.append(quantity_4)           
        label = tk.Label(window, textvariable=quantities_4[row]).grid(column=5, row=row+100)
        # Column 6 showing 5th selected quantities
        quantity_5 = tk.StringVar()
        quantity_5.set("                         ")
        quantities_5.append(quantity_5)           
        label = tk.Label(window, textvariable=quantities_5[row]).grid(column=6, row=row+100)
        # Column 7 showing 6th selected quantities
        quantity_6 = tk.StringVar()
        quantity_6.set("                         ")
        quantities_6.append(quantity_6)           
        label = tk.Label(window, textvariable=quantities_6[row]).grid(column=7, row=row+100)

    # Create the Select button
    button = ttk.Button(text="Select", command=display_selection)
    button.place(x=1370, y=23)

    window.mainloop()



def function_2():
    ...


def function_3():
    ...


if __name__ == "__main__":
    main()