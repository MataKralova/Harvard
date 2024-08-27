from openpyxl import load_workbook
from operator import itemgetter
import os
import tkinter as tk
from tkinter import ttk


def main():
    file_path, wb, ws, keys, data = process_excel()
    #create_widget(keys)

    ####################### WIDGET #####################################

    #def create_widget(keys):
    # Create tkinter window
    window = tk.Tk(className=" Food Nutritional Values") 
    window.geometry("1500x555")

    # Define label for the nutrients lists
    nutrients_label = tk.Label(window, text=" Select nutrients:").grid(column=1, row=15, padx=10, pady=25)

    # Define label for optimal quantities
    optima_label = tk.Label(window, text=" Edit optimal quantities:").grid(column=1, row=16, padx=10, pady=(0, 0))

    # Create display values for comboboxes for nutrients
    n1 = tk.StringVar()
    n2 = tk.StringVar()
    n3 = tk.StringVar()
    n4 = tk.StringVar()
    n5 = tk.StringVar()
    n6 = tk.StringVar()
    
    # Create display values for entry fields for optimal quantities
    o1 = tk.StringVar()
    o2 = tk.StringVar()
    o3 = tk.StringVar()
    o4 = tk.StringVar()
    o5 = tk.StringVar()
    o6 = tk.StringVar()

    # Create comboboxes for nutrients
    nutrients_1 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n1)
    nutrients_2 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n2)
    nutrients_3 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n3)
    nutrients_4 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n4)
    nutrients_5 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n5)
    nutrients_6 = ttk.Combobox(window, state="readonly", justify="center", width=24, textvariable=n6)

    # Create entry fields for optimal quantities
    optimum_1 = tk.Entry(window, justify="center", width=24, textvariable=o1)
    optimum_2 = tk.Entry(window, justify="center", width=24, textvariable=o2)
    optimum_3 = tk.Entry(window, justify="center", width=24, textvariable=o3)
    optimum_4 = tk.Entry(window, justify="center", width=24, textvariable=o4)
    optimum_5 = tk.Entry(window, justify="center", width=24, textvariable=o5)
    optimum_6 = tk.Entry(window, justify="center", width=24, textvariable=o6)

    # Add combobox dropdown lists
    values = []
    for key in keys:
        values.append(key)
    nutrients_1["values"] = values
    nutrients_2["values"] = values
    nutrients_3["values"] = values
    nutrients_4["values"] = values
    nutrients_5["values"] = values
    nutrients_6["values"] = values

    # Show comboboxes with default values
    nutrients_1.grid(column=2, row=15)
    nutrients_1.current(0)
    nutrients_2.grid(column=3, row=15)
    nutrients_2.current(0)
    nutrients_3.grid(column=4, row=15)
    nutrients_3.current(0)
    nutrients_4.grid(column=5, row=15)
    nutrients_4.current(0)
    nutrients_5.grid(column=6, row=15)
    nutrients_5.current(0)
    nutrients_6.grid(column=7, row=15)
    nutrients_6.current(0)

    # Show entry fields with default values
    optimum_1.grid(column=2, row=16)
    optimum_2.grid(column=3, row=16)
    optimum_3.grid(column=4, row=16)
    optimum_4.grid(column=5, row=16)
    optimum_5.grid(column=6, row=16)
    optimum_6.grid(column=7, row=16)

    # Create StringVar lists to associate with labels
    foods = []
    sizes = []
    quantities_1 = []
    quantities_2 = []
    quantities_3 = []
    quantities_4 = []
    quantities_5 = []
    quantities_6 = []

    # Create result labels
    for row in range (0,10):
        # Column 0 showing foods
        food = tk.StringVar()
        food.set("                              ")
        foods.append(food)
        label = tk.Label(window, textvariable=foods[row]).grid(column=0, row=row+100, padx=25, pady=(20, 0))
        # Column 1 showing sizes
        size = tk.StringVar()
        size.set("               ")
        sizes.append(size)
        label = tk.Label(window, textvariable=sizes[row]).grid(column=1, row=row+100, pady=(20, 0))
        # Column 2 showing main selected quantities
        quantity_1 = tk.StringVar()
        quantity_1.set("                         ")
        quantities_1.append(quantity_1)           
        label = tk.Label(window, textvariable=quantities_1[row]).grid(column=2, row=row+100, pady=(20, 0))
        # Column 3 showing 2nd selected quantities
        quantity_2 = tk.StringVar()
        quantity_2.set("                         ")
        quantities_2.append(quantity_2)           
        label = tk.Label(window, textvariable=quantities_2[row]).grid(column=3, row=row+100, pady=(20, 0))
        # Column 4 showing 3rd selected quantities
        quantity_3 = tk.StringVar()
        quantity_3.set("                         ")
        quantities_3.append(quantity_3)           
        label = tk.Label(window, textvariable=quantities_3[row]).grid(column=4, row=row+100, pady=(20, 0))
        # Column 5 showing 4th selected quantities
        quantity_4 = tk.StringVar()
        quantity_4.set("                         ")
        quantities_4.append(quantity_4)           
        label = tk.Label(window, textvariable=quantities_4[row]).grid(column=5, row=row+100, pady=(20, 0))
        # Column 6 showing 5th selected quantities
        quantity_5 = tk.StringVar()
        quantity_5.set("                         ")
        quantities_5.append(quantity_5)           
        label = tk.Label(window, textvariable=quantities_5[row]).grid(column=6, row=row+100, pady=(20, 0))
        # Column 7 showing 6th selected quantities
        quantity_6 = tk.StringVar()
        quantity_6.set("                         ")
        quantities_6.append(quantity_6)           
        label = tk.Label(window, textvariable=quantities_6[row]).grid(column=7, row=row+100, pady=(20, 0))


    def display_selection():
        # Create a list sorted from biggest by values in main selected column
        new_list = sorted(data[0:175], key=itemgetter(nutrients_1.get()), reverse=True)
        
        # Populate result values for column "Potravina (100 g)"
        for row in range (0,10):
            foods[row].set(new_list[row]["Potravina (100 g)"])

        # Populate result values for column "Jedlo (obvyklá porcia)"
        for row in range (0,10):
            sizes[row].set(new_list[row]["Jedlo (obvyklá porcia)"])

        # Populate result values for main selected quantities column
        for row in range (0,10):
            quantities_1[row].set(new_list[row][nutrients_1.get()])

        # Populate result values for 2nd selected quantities column
        for row in range (0,10):
            quantities_2[row].set(new_list[row][nutrients_2.get()])

        # Populate result values for 3rd selected quantities column
        for row in range (0,10):
            quantities_3[row].set(new_list[row][nutrients_3.get()])

        # Populate result values for 4th selected quantities column
        for row in range (0,10):
            quantities_4[row].set(new_list[row][nutrients_4.get()])

        # Populate result values for 5th selected quantities column
        for row in range (0,10):
            quantities_5[row].set(new_list[row][nutrients_5.get()])

        # Populate result values for 6th selected quantities column
        for row in range (0,10):
            quantities_6[row].set(new_list[row][nutrients_6.get()])

        # Show optimal quantities, if non-existent, show blank
        if data[175][nutrients_1.get()]:
            o1.set(data[175][nutrients_1.get()])
        else:
            o1.set("")
        if data[175][nutrients_2.get()]:
            o2.set(data[175][nutrients_2.get()])
        else:
            o2.set("")
        if data[175][nutrients_3.get()]:
            o3.set(data[175][nutrients_3.get()])
        else:
            o3.set("")
        if data[175][nutrients_4.get()]:
            o4.set(data[175][nutrients_4.get()])
        else:
            o4.set("")
        if data[175][nutrients_5.get()]:
            o5.set(data[175][nutrients_5.get()])
        else:
            o5.set("")
        if data[175][nutrients_6.get()]:
            o6.set(data[175][nutrients_6.get()])
        else:
            o6.set("")


    def save():
        # Write optimal quantities into data (in data, row 175 contains the optimal quantities)
        print(data[175][nutrients_1.get()])
        entry_1 = optimum_1.get()
        data[175][nutrients_1.get()] = entry_1
        print(data[175][nutrients_1.get()])
        entry_2 = optimum_2.get()
        data[175][nutrients_2.get()] = entry_2
        entry_3 = optimum_3.get()
        data[175][nutrients_3.get()] = entry_3
        entry_4 = optimum_4.get()
        data[175][nutrients_4.get()] = entry_4
        entry_5 = optimum_5.get()
        data[175][nutrients_5.get()] = entry_5
        entry_6 = optimum_6.get()
        data[175][nutrients_6.get()] = entry_6

        # Write optimal quantities into Excel (in Excel, row 177 contains the optimal quantities)
        column = keys.index(nutrients_1.get()) + 1
        c1 = ws.cell(row=177, column=column)
        c1.value = entry_1
        column = keys.index(nutrients_2.get()) + 1
        c2 = ws.cell(row=177, column=column)
        c2.value = entry_2
        column = keys.index(nutrients_3.get()) + 1
        c3 = ws.cell(row=177, column=column)
        c3.value = entry_3
        column = keys.index(nutrients_4.get()) + 1
        c4 = ws.cell(row=177, column=column)
        c4.value = entry_4
        column = keys.index(nutrients_5.get()) + 1
        c5 = ws.cell(row=177, column=column)
        c5.value = entry_5
        column = keys.index(nutrients_6.get()) + 1
        c6 = ws.cell(row=177, column=column)
        c6.value = entry_6

        # Save Excel
        wb.save(file_path)


    # Create the Select button
    button = ttk.Button(text="Select", command=display_selection)
    button.place(x=1370, y=23)

    # Create the Save button
    button = ttk.Button(text="Save", command=save)
    button.place(x=1370, y=69)

    window.mainloop()


def process_excel():
    # Get the path to the Excel file
    base_path = getattr(os.sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(base_path, "Food Data.xlsx")

    # Load the entire workbook
    wb = load_workbook(file_path)

    ws = wb["Values"]
    all_rows = list(ws.rows)

    # Get the first row as keys
    keys = [cell.value for cell in ws[1]]

    # Create a list of dictionaries
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(keys, row))
        data.append(row_dict)
    
    return file_path, wb, ws, keys, data


if __name__ == "__main__":
    main()